import pandas as pd
from datetime import datetime, date, timezone, timedelta
from typing import Dict, List, Optional
from config.firebaseConfig import firebase_config
from config.settings import Settings
from utils.logger import app_logger, log_audit
from services.firebaseWrapper import requires_connection
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class ReportService:
    """Servicio para generar y exportar reportes"""
    
    # Constantes
    MAX_HISTORIAL_REPORTES = 100
    PREVIEW_LIMIT = 50
    
    def __init__(self):
        try:
            self.db = firebase_config.get_firestore_client()
            self.datos_ref = self.db.collection(Settings.COLLECTION_DATOS_TRIBUTARIOS)
            self.reportes_ref = self.db.collection(Settings.COLLECTION_REPORTES)
            self.usuarios_ref = self.db.collection(Settings.COLLECTION_USUARIOS)
            self.chile_tz = timezone(timedelta(hours=-3))
            
            # Cache para RUTs de clientes
            self._rut_cache = {}
            
            app_logger.info("ReportService inicializado correctamente")
        except Exception as e:
            app_logger.error(f"Error al inicializar ReportService: {str(e)}")
            raise
    
    def get_chile_time(self) -> datetime:
        """
        Retorna la fecha/hora actual en zona horaria de Chile
        
        Returns:
            datetime: Fecha/hora actual en Chile (UTC-3)
        """
        return datetime.now(self.chile_tz)
    
    def _validar_filtros(self, filtros: Dict) -> tuple[bool, Optional[str]]:
        """
        Valida que los filtros sean correctos
        
        Args:
            filtros (Dict): Diccionario con filtros a validar
            
        Returns:
            tuple[bool, Optional[str]]: (es_valido, mensaje_error)
        """
        try:
            # Validar fechas
            if "fecha_desde" in filtros and "fecha_hasta" in filtros:
                fecha_desde = filtros["fecha_desde"]
                fecha_hasta = filtros["fecha_hasta"]
                
                if fecha_desde and fecha_hasta and fecha_desde > fecha_hasta:
                    return False, "La fecha 'desde' no puede ser mayor que la fecha 'hasta'"
            
            # Validar tipos de impuesto permitidos
            tipos_validos = ["IVA", "Renta", "Importación", "Exportación", "Otro"]
            if "tipo_impuesto" in filtros and filtros["tipo_impuesto"]:
                if filtros["tipo_impuesto"] not in tipos_validos:
                    return False, f"Tipo de impuesto inválido. Debe ser uno de: {', '.join(tipos_validos)}"
            
            # Validar países permitidos
            paises_validos = ["Chile", "Perú", "Colombia"]
            if "pais" in filtros and filtros["pais"]:
                if filtros["pais"] not in paises_validos:
                    return False, f"País inválido. Debe ser uno de: {', '.join(paises_validos)}"
            
            # Validar estado
            estados_validos = ["local", "bolsa", "ambos"]
            if "estado" in filtros and filtros["estado"]:
                if filtros["estado"] not in estados_validos:
                    return False, f"Estado inválido. Debe ser uno de: {', '.join(estados_validos)}"
            
            return True, None
            
        except Exception as e:
            app_logger.error(f"Error al validar filtros: {str(e)}")
            return False, "Error interno al validar filtros"
    
    @requires_connection
    def obtener_datos_filtrados(self, filtros: Dict, usuario_id: str, rol: str) -> List[Dict]:
        """
        Obtiene calificaciones tributarias según filtros
        
        Args:
            filtros (Dict): Filtros a aplicar
            usuario_id (str): ID del usuario que genera el reporte
            rol (str): Rol del usuario
            
        Returns:
            List[Dict]: Lista de calificaciones
        """
        try:
            # Validar filtros primero
            es_valido, mensaje_error = self._validar_filtros(filtros)
            if not es_valido:
                app_logger.warning(f"Filtros inválidos: {mensaje_error}")
                return []
            
            app_logger.info(f"Obteniendo datos filtrados para usuario {usuario_id} con rol {rol}")
            
            # Query base optimizada
            query = self.datos_ref.where("activo", "==", True)
            
            # Ejecutar query base
            docs = query.stream()
            
            calificaciones = []
            registros_procesados = 0
            registros_filtrados = 0
            
            for doc in docs:
                registros_procesados += 1
                data = doc.to_dict()
                data["_id"] = doc.id
                
                # Filtro de fechas
                if "fecha_desde" in filtros and filtros["fecha_desde"]:
                    fecha_str = filtros["fecha_desde"].strftime("%Y-%m-%d")
                    if data.get("fechaDeclaracion", "") < fecha_str:
                        registros_filtrados += 1
                        continue
                
                if "fecha_hasta" in filtros and filtros["fecha_hasta"]:
                    fecha_str = filtros["fecha_hasta"].strftime("%Y-%m-%d")
                    if data.get("fechaDeclaracion", "") > fecha_str:
                        registros_filtrados += 1
                        continue
                
                # Filtro de tipo impuesto
                if "tipo_impuesto" in filtros and filtros["tipo_impuesto"]:
                    if data.get("tipoImpuesto", "") != filtros["tipo_impuesto"]:
                        registros_filtrados += 1
                        continue
                
                # Filtro de país
                if "pais" in filtros and filtros["pais"]:
                    if data.get("pais", "") != filtros["pais"]:
                        registros_filtrados += 1
                        continue
                
                # Filtrar por estado (local/bolsa)
                es_local = data.get("esLocal", False)
                estado_filtro = filtros.get("estado", "ambos")
                
                if estado_filtro == "local" and not es_local:
                    registros_filtrados += 1
                    continue
                if estado_filtro == "bolsa" and es_local:
                    registros_filtrados += 1
                    continue
                
                # Permisos según rol
                es_propietario = data.get("propietarioRegistroId") == usuario_id
                
                if rol == "administrador":
                    calificaciones.append(data)
                elif rol == "auditor_tributario":
                    calificaciones.append(data)
                elif not es_local:
                    calificaciones.append(data)
                elif es_propietario:
                    calificaciones.append(data)
                else:
                    registros_filtrados += 1
            
            app_logger.info(
                f"Datos filtrados: {len(calificaciones)} registros obtenidos "
                f"de {registros_procesados} procesados ({registros_filtrados} filtrados)"
            )
            return calificaciones
        
        except Exception as e:
            app_logger.error(f"Error al obtener datos filtrados: {str(e)}", exc_info=True)
            return []
    
    def obtener_rut_cliente(self, cliente_id) -> str:
        """
        Obtiene el RUT de un cliente por su ID (con caché)
        
        Args:
            cliente_id: ID del cliente (puede ser str, list, o None)
            
        Returns:
            str: RUT del cliente o "N/A"
        """
        # Manejar diferentes tipos de cliente_id
        if not cliente_id:
            return "N/A"
        
        # Si es una lista o tupla, tomar el primer elemento
        if isinstance(cliente_id, (list, tuple)):
            if len(cliente_id) > 0:
                cliente_id = cliente_id[0]
            else:
                return "N/A"
        
        # Convertir a string
        cliente_id = str(cliente_id)
        
        # Verificar que no esté vacío
        if not cliente_id or cliente_id == "None":
            return "N/A"
        
        # Verificar caché primero
        if cliente_id in self._rut_cache:
            return self._rut_cache[cliente_id]
        
        try:
            doc = self.usuarios_ref.document(cliente_id).get()
            if doc.exists:
                rut = doc.to_dict().get("rut", "N/A")
                self._rut_cache[cliente_id] = rut
                return rut
            else:
                app_logger.warning(f"Cliente no encontrado: {cliente_id}")
                self._rut_cache[cliente_id] = "N/A"
                return "N/A"
                
        except Exception as e:
            app_logger.error(f"Error al obtener RUT del cliente {cliente_id}: {str(e)}")
            return "N/A"
    
    def limpiar_cache_rut(self):
        """Limpia el caché de RUTs de clientes"""
        self._rut_cache.clear()
        app_logger.debug("Caché de RUTs limpiado")
    
    def preparar_dataframe(self, calificaciones: List[Dict]) -> pd.DataFrame:
        """
        Convierte las calificaciones a DataFrame para exportación
        
        Args:
            calificaciones (List[Dict]): Lista de calificaciones
            
        Returns:
            pd.DataFrame: DataFrame con los datos formateados
        """
        if not calificaciones:
            return pd.DataFrame()
        
        try:
            data = []
            for cal in calificaciones:
                # Obtener RUT del cliente - manejar diferentes tipos
                cliente_id = cal.get("clienteId", "")
                
                # Si clienteId es una lista, tomar el primer elemento
                if isinstance(cliente_id, (list, tuple)):
                    cliente_id = cliente_id[0] if len(cliente_id) > 0 else ""
                
                # Convertir a string
                cliente_id = str(cliente_id) if cliente_id else ""
                
                # Obtener RUT
                if cliente_id and cliente_id != "None":
                    rut_cliente = self.obtener_rut_cliente(cliente_id)
                else:
                    rut_cliente = "N/A"
                
                # Calcular suma factores 8-19
                factores = cal.get("factores", {})
                
                # Validar que factores sea un diccionario
                if not isinstance(factores, dict):
                    app_logger.warning(f"Factores no es diccionario: {type(factores)}")
                    factores = {}
                
                suma_8_19 = 0
                for i in range(8, 20):
                    factor_val = factores.get(f"factor_{i}", 0)
                    
                    # Manejar listas, tuplas u otros tipos
                    if isinstance(factor_val, (list, tuple)):
                        factor_val = factor_val[0] if len(factor_val) > 0 else 0
                    
                    # Convertir a numérico
                    if isinstance(factor_val, (int, float)):
                        suma_8_19 += factor_val
                    else:
                        try:
                            suma_8_19 += float(factor_val)
                        except (ValueError, TypeError):
                            app_logger.warning(f"Factor {i} no numérico: {factor_val}")
                            continue
                
                # Fila base
                fila = {
                    "RUT Cliente": str(rut_cliente),
                    "Fecha Declaración": str(cal.get("fechaDeclaracion", "")),
                    "Tipo Impuesto": str(cal.get("tipoImpuesto", "")),
                    "País": str(cal.get("pais", "")),
                    "Monto Declarado": cal.get("montoDeclarado", 0) if isinstance(cal.get("montoDeclarado", 0), (int, float)) else 0,
                }
                
                # Agregar factores 1-19
                for i in range(1, 20):
                    factor_value = factores.get(f"factor_{i}", 0)
                    
                    # Convertir listas, tuplas u otros tipos incompatibles
                    if isinstance(factor_value, (list, tuple)):
                        factor_value = factor_value[0] if len(factor_value) > 0 else 0
                    elif isinstance(factor_value, dict):
                        factor_value = 0
                    
                    # Asegurar que sea numérico
                    if not isinstance(factor_value, (int, float)):
                        try:
                            factor_value = float(factor_value)
                        except (ValueError, TypeError):
                            factor_value = 0
                    
                    fila[f"Factor {i}"] = factor_value
                
                # Agregar suma y metadatos
                fila["Suma Factores 8-19"] = round(suma_8_19, 4)
                fila["Estado"] = "Local" if cal.get("esLocal", False) else "Bolsa"
                fila["Válido"] = "Sí" if suma_8_19 <= 1.0 else "No (>1.0)"
                
                # Agregar campos adicionales (asegurar que sean strings)
                propietario_id = cal.get("propietarioRegistroId", "N/A")
                if isinstance(propietario_id, str) and len(propietario_id) > 0:
                    fila["Usuario Registrador"] = propietario_id[:8]
                else:
                    fila["Usuario Registrador"] = "N/A"
                
                fecha_registro = cal.get("fechaRegistro", "N/A")
                fila["Fecha Registro"] = str(fecha_registro)
                
                data.append(fila)
            
            df = pd.DataFrame(data)
            app_logger.info(f"DataFrame preparado con {len(df)} registros")
            return df
            
        except Exception as e:
            app_logger.error(f"Error al preparar DataFrame: {str(e)}", exc_info=True)
            return pd.DataFrame()
    
    def exportar_csv(self, file_path: str, calificaciones: List[Dict], 
                    filtros: Dict, usuario_id: str) -> Dict:
        """
        Exporta calificaciones a CSV
        
        Args:
            file_path (str): Ruta donde guardar el archivo
            calificaciones (List[Dict]): Datos a exportar
            filtros (Dict): Filtros aplicados
            usuario_id (str): ID del usuario
            
        Returns:
            Dict: Resultado de la exportación
        """
        try:
            if not calificaciones:
                return {
                    "success": False,
                    "message": "No hay datos para exportar"
                }
            
            app_logger.info(f"Iniciando exportación CSV para usuario {usuario_id}")
            
            # Crear DataFrame
            df = self.preparar_dataframe(calificaciones)
            
            if df.empty:
                return {
                    "success": False,
                    "message": "Error al preparar los datos para exportación"
                }
            
            # Exportar a CSV con codificación UTF-8 BOM (compatible con Excel)
            df.to_csv(file_path, index=False, encoding='utf-8-sig', sep=',')
            
            # Obtener nombre del archivo
            nombre_archivo = file_path.split("/")[-1].split("\\")[-1]
            
            # Registrar en Firebase
            registro_exitoso = self.registrar_reporte(
                usuario_id=usuario_id,
                tipo_reporte="exportacion_calificaciones",
                filtros=filtros,
                total_registros=len(calificaciones),
                formato="CSV",
                nombre_archivo=nombre_archivo
            )
            
            if not registro_exitoso:
                app_logger.warning("CSV exportado pero no se pudo registrar en Firebase")
            
            app_logger.info(f"CSV exportado exitosamente: {file_path}")
            
            return {
                "success": True,
                "message": f"CSV generado exitosamente con {len(calificaciones):,} registros.\nArchivo guardado en: {nombre_archivo}",
                "file_path": file_path
            }
        
        except PermissionError:
            error_msg = "No se puede guardar el archivo. Verifique que no esté abierto en otra aplicación."
            app_logger.error(f"Error de permisos al exportar CSV: {error_msg}")
            return {
                "success": False,
                "message": error_msg
            }
        except Exception as e:
            app_logger.error(f"Error al exportar CSV: {str(e)}", exc_info=True)
            return {
                "success": False,
                "message": f"Error al exportar: {str(e)}"
            }
    
    def exportar_excel(self, file_path: str, calificaciones: List[Dict], 
                      filtros: Dict, usuario_id: str) -> Dict:
        """
        Exporta calificaciones a Excel con formato profesional
        
        Args:
            file_path (str): Ruta donde guardar el archivo
            calificaciones (List[Dict]): Datos a exportar
            filtros (Dict): Filtros aplicados
            usuario_id (str): ID del usuario
            
        Returns:
            Dict: Resultado de la exportación
        """
        try:
            if not calificaciones:
                return {
                    "success": False,
                    "message": "No hay datos para exportar"
                }
            
            app_logger.info(f"Iniciando exportación Excel para usuario {usuario_id}")
            
            # Crear DataFrame
            df = self.preparar_dataframe(calificaciones)
            
            if df.empty:
                return {
                    "success": False,
                    "message": "Error al preparar los datos para exportación"
                }
            
            # Crear workbook
            wb = Workbook()
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="E94E1B", end_color="E94E1B", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            border_thin = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # ========== HOJA 1: Datos ==========
            ws_datos = wb.active
            ws_datos.title = "Calificaciones Tributarias"
            
            # Escribir datos
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_datos.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = border_thin
                    
                    # Formato de encabezado
                    if r_idx == 1:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                    else:
                        # Alineación según tipo de dato
                        if c_idx <= 4:  # Texto
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                        else:  # Números
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                        
                        # Formato de número para montos
                        if c_idx == 5 and isinstance(value, (int, float)):  # Monto Declarado
                            cell.number_format = '$#,##0.00'
                        elif c_idx >= 6 and c_idx <= 24 and isinstance(value, (int, float)):  # Factores
                            cell.number_format = '0.0000'
            
            # Ajustar anchos de columna
            for column in ws_datos.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                adjusted_width = min(max_length + 3, 50)
                ws_datos.column_dimensions[column_letter].width = adjusted_width
            
            # Congelar primera fila
            ws_datos.freeze_panes = "A2"
            
            # ========== HOJA 2: Resumen ==========
            ws_resumen = wb.create_sheet("Resumen Ejecutivo")
            
            # Título principal
            ws_resumen['A1'] = "REPORTE DE CALIFICACIONES TRIBUTARIAS"
            ws_resumen['A1'].font = Font(size=16, bold=True, color="E94E1B")
            ws_resumen.merge_cells('A1:B1')
            
            row_actual = 3
            
            # Metadatos
            ws_resumen[f'A{row_actual}'] = "📅 Fecha de generación:"
            ws_resumen[f'B{row_actual}'] = self.get_chile_time().strftime("%d/%m/%Y %H:%M:%S")
            ws_resumen[f'A{row_actual}'].font = Font(bold=True)
            row_actual += 1
            
            ws_resumen[f'A{row_actual}'] = "👤 Usuario generador:"
            ws_resumen[f'B{row_actual}'] = usuario_id[:8] + "..."
            ws_resumen[f'A{row_actual}'].font = Font(bold=True)
            row_actual += 1
            
            ws_resumen[f'A{row_actual}'] = "📊 Total de registros:"
            ws_resumen[f'B{row_actual}'] = len(calificaciones)
            ws_resumen[f'A{row_actual}'].font = Font(bold=True)
            ws_resumen[f'B{row_actual}'].font = Font(bold=True, color="E94E1B", size=12)
            row_actual += 2
            
            # Filtros aplicados
            ws_resumen[f'A{row_actual}'] = "🔍 FILTROS APLICADOS"
            ws_resumen[f'A{row_actual}'].font = Font(size=12, bold=True, color="2c3e50")
            row_actual += 1
            
            if "fecha_desde" in filtros and filtros["fecha_desde"]:
                ws_resumen[f'A{row_actual}'] = "Fecha desde:"
                ws_resumen[f'B{row_actual}'] = filtros["fecha_desde"].strftime("%d/%m/%Y")
                row_actual += 1
            
            if "fecha_hasta" in filtros and filtros["fecha_hasta"]:
                ws_resumen[f'A{row_actual}'] = "Fecha hasta:"
                ws_resumen[f'B{row_actual}'] = filtros["fecha_hasta"].strftime("%d/%m/%Y")
                row_actual += 1
            
            if "tipo_impuesto" in filtros and filtros["tipo_impuesto"]:
                ws_resumen[f'A{row_actual}'] = "Tipo de impuesto:"
                ws_resumen[f'B{row_actual}'] = filtros["tipo_impuesto"]
                row_actual += 1
            
            if "pais" in filtros and filtros["pais"]:
                ws_resumen[f'A{row_actual}'] = "País:"
                ws_resumen[f'B{row_actual}'] = filtros["pais"]
                row_actual += 1
            
            if "estado" in filtros and filtros["estado"]:
                ws_resumen[f'A{row_actual}'] = "Estado:"
                ws_resumen[f'B{row_actual}'] = filtros["estado"].upper()
                row_actual += 1
            
            row_actual += 1
            
            # Estadísticas
            ws_resumen[f'A{row_actual}'] = "📈 ESTADÍSTICAS"
            ws_resumen[f'A{row_actual}'].font = Font(size=12, bold=True, color="2c3e50")
            row_actual += 1
            
            # Calcular estadísticas
            locales = sum(1 for c in calificaciones if c.get("esLocal", False))
            bolsa = len(calificaciones) - locales
            total_monto = sum(c.get("montoDeclarado", 0) for c in calificaciones)
            
            # Calcular registros válidos/inválidos
            registros_validos = 0
            registros_invalidos = 0
            for c in calificaciones:
                factores = c.get("factores", {})
                suma_8_19 = sum(factores.get(f"factor_{i}", 0) for i in range(8, 20))
                if suma_8_19 <= 1.0:
                    registros_validos += 1
                else:
                    registros_invalidos += 1
            
            ws_resumen[f'A{row_actual}'] = "📋 Registros Locales:"
            ws_resumen[f'B{row_actual}'] = locales
            ws_resumen[f'A{row_actual}'].font = Font(bold=True)
            row_actual += 1
            
            ws_resumen[f'A{row_actual}'] = "🏛️ Registros Bolsa:"
            ws_resumen[f'B{row_actual}'] = bolsa
            ws_resumen[f'A{row_actual}'].font = Font(bold=True)
            row_actual += 1
            
            ws_resumen[f'A{row_actual}'] = "✅ Registros Válidos:"
            ws_resumen[f'B{row_actual}'] = registros_validos
            ws_resumen[f'A{row_actual}'].font = Font(bold=True)
            ws_resumen[f'B{row_actual}'].font = Font(color="27ae60", bold=True)
            row_actual += 1
            
            ws_resumen[f'A{row_actual}'] = "❌ Registros Inválidos:"
            ws_resumen[f'B{row_actual}'] = registros_invalidos
            ws_resumen[f'A{row_actual}'].font = Font(bold=True)
            if registros_invalidos > 0:
                ws_resumen[f'B{row_actual}'].font = Font(color="e74c3c", bold=True)
            row_actual += 1
            
            ws_resumen[f'A{row_actual}'] = "💰 Monto Total Declarado:"
            ws_resumen[f'B{row_actual}'] = total_monto
            ws_resumen[f'B{row_actual}'].number_format = '$#,##0.00'
            ws_resumen[f'A{row_actual}'].font = Font(bold=True)
            ws_resumen[f'B{row_actual}'].font = Font(bold=True, size=12)
            row_actual += 2
            
            # Porcentajes
            if len(calificaciones) > 0:
                ws_resumen[f'A{row_actual}'] = "📊 DISTRIBUCIÓN PORCENTUAL"
                ws_resumen[f'A{row_actual}'].font = Font(size=12, bold=True, color="2c3e50")
                row_actual += 1
                
                pct_locales = (locales / len(calificaciones)) * 100
                pct_bolsa = (bolsa / len(calificaciones)) * 100
                pct_validos = (registros_validos / len(calificaciones)) * 100
                
                ws_resumen[f'A{row_actual}'] = "% Locales:"
                ws_resumen[f'B{row_actual}'] = f"{pct_locales:.1f}%"
                row_actual += 1
                
                ws_resumen[f'A{row_actual}'] = "% Bolsa:"
                ws_resumen[f'B{row_actual}'] = f"{pct_bolsa:.1f}%"
                row_actual += 1
                
                ws_resumen[f'A{row_actual}'] = "% Validez:"
                ws_resumen[f'B{row_actual}'] = f"{pct_validos:.1f}%"
                if pct_validos < 90:
                    ws_resumen[f'B{row_actual}'].font = Font(color="e74c3c", bold=True)
                else:
                    ws_resumen[f'B{row_actual}'].font = Font(color="27ae60", bold=True)
            
            # Ajustar anchos de la hoja resumen
            ws_resumen.column_dimensions['A'].width = 35
            ws_resumen.column_dimensions['B'].width = 45
            
            # Guardar archivo
            wb.save(file_path)
            
            # Obtener nombre del archivo
            nombre_archivo = file_path.split("/")[-1].split("\\")[-1]
            
            # Registrar en Firebase
            registro_exitoso = self.registrar_reporte(
                usuario_id=usuario_id,
                tipo_reporte="exportacion_calificaciones",
                filtros=filtros,
                total_registros=len(calificaciones),
                formato="Excel",
                nombre_archivo=nombre_archivo
            )
            
            if not registro_exitoso:
                app_logger.warning("Excel exportado pero no se pudo registrar en Firebase")
            
            app_logger.info(f"Excel exportado exitosamente: {file_path}")
            
            return {
                "success": True,
                "message": f"Excel generado exitosamente con {len(calificaciones):,} registros.\nArchivo guardado en: {nombre_archivo}",
                "file_path": file_path
            }
        
        except PermissionError:
            error_msg = "No se puede guardar el archivo. Verifique que no esté abierto en otra aplicación."
            app_logger.error(f"Error de permisos al exportar Excel: {error_msg}")
            return {
                "success": False,
                "message": error_msg
            }
        except Exception as e:
            app_logger.error(f"Error al exportar Excel: {str(e)}", exc_info=True)
            return {
                "success": False,
                "message": f"Error al exportar: {str(e)}"
            }
    
    @requires_connection
    def registrar_reporte(self, usuario_id: str, tipo_reporte: str, 
                         filtros: Dict, total_registros: int, 
                         formato: str, nombre_archivo: str) -> bool:
        """
        Registra un reporte generado en Firebase
        
        Args:
            usuario_id (str): ID del usuario que generó el reporte
            tipo_reporte (str): Tipo de reporte
            filtros (Dict): Filtros aplicados
            total_registros (int): Total de registros exportados
            formato (str): Formato del reporte (CSV/Excel)
            nombre_archivo (str): Nombre del archivo generado
            
        Returns:
            bool: True si se registró correctamente
        """
        try:
            # Convertir datetime.date a datetime para Firestore
            filtros_firestore = {}
            for key, value in filtros.items():
                if isinstance(value, date) and not isinstance(value, datetime):
                    # Convertir date a datetime
                    filtros_firestore[key] = datetime.combine(value, datetime.min.time())
                else:
                    filtros_firestore[key] = value
            
            # Crear diccionario del reporte
            reporte_data = {
                "usuarioGeneradorId": usuario_id,
                "tipoReporte": tipo_reporte,
                "filtrosAplicados": filtros_firestore,
                "totalRegistros": total_registros,
                "formato": formato,
                "nombreArchivo": nombre_archivo,
                "fechaGeneracion": self.get_chile_time()
            }
            
            # Guardar en Firestore
            doc_ref = self.reportes_ref.add(reporte_data)
            
            # Registrar auditoría
            log_audit(
                action="REPORTE_GENERADO",
                user_id=usuario_id,
                details={
                    "tipo": tipo_reporte,
                    "formato": formato,
                    "registros": total_registros,
                    "archivo": nombre_archivo,
                    "reporte_id": doc_ref[1].id
                }
            )
            
            app_logger.info(f"Reporte registrado exitosamente: {nombre_archivo} (ID: {doc_ref[1].id})")
            return True
        
        except Exception as e:
            app_logger.error(f"Error al registrar reporte: {str(e)}", exc_info=True)
            return False
    
    @requires_connection
    def obtener_historial_reportes(self, usuario_id: str, rol: str, limite: int = None) -> List[Dict]:
        """
        Obtiene el historial de reportes generados
        
        Args:
            usuario_id (str): ID del usuario
            rol (str): Rol del usuario
            limite (int, optional): Cantidad máxima de reportes a obtener
            
        Returns:
            List[Dict]: Lista de reportes
        """
        try:
            if limite is None:
                limite = self.MAX_HISTORIAL_REPORTES
            
            app_logger.info(f"Obteniendo historial de reportes para {usuario_id} (rol: {rol})")
            
            # Admin y auditores ven todos, otros usuarios ven solo los suyos
            if rol in ["administrador", "auditor_tributario"]:
                query = self.reportes_ref\
                    .order_by("fechaGeneracion", direction="DESCENDING")\
                    .limit(limite)
            else:
                query = self.reportes_ref\
                    .where("usuarioGeneradorId", "==", usuario_id)\
                    .order_by("fechaGeneracion", direction="DESCENDING")\
                    .limit(limite)
            
            docs = query.stream()
            
            reportes = []
            for doc in docs:
                data = doc.to_dict()
                data["_id"] = doc.id
                reportes.append(data)
            
            app_logger.info(f"Se encontraron {len(reportes)} reportes en el historial")
            return reportes
        
        except Exception as e:
            app_logger.error(f"Error al obtener historial de reportes: {str(e)}", exc_info=True)
            return []
    
    def obtener_estadisticas_reportes(self, usuario_id: str, rol: str) -> Dict:
        """
        Obtiene estadísticas de reportes generados
        
        Args:
            usuario_id (str): ID del usuario
            rol (str): Rol del usuario
            
        Returns:
            Dict: Estadísticas de reportes
        """
        try:
            reportes = self.obtener_historial_reportes(usuario_id, rol, limite=500)
            
            if not reportes:
                return {
                    "total_reportes": 0,
                    "reportes_csv": 0,
                    "reportes_excel": 0,
                    "total_registros_exportados": 0
                }
            
            csv_count = sum(1 for r in reportes if r.get("formato") == "CSV")
            excel_count = sum(1 for r in reportes if r.get("formato") == "Excel")
            total_registros = sum(r.get("totalRegistros", 0) for r in reportes)
            
            return {
                "total_reportes": len(reportes),
                "reportes_csv": csv_count,
                "reportes_excel": excel_count,
                "total_registros_exportados": total_registros,
                "ultimo_reporte": reportes[0].get("fechaGeneracion") if reportes else None
            }
            
        except Exception as e:
            app_logger.error(f"Error al obtener estadísticas de reportes: {str(e)}")
            return {
                "total_reportes": 0,
                "reportes_csv": 0,
                "reportes_excel": 0,
                "total_registros_exportados": 0
            }